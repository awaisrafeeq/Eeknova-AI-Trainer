#!/usr/bin/env python3
"""Convert the 18-step Zumba choreography workbook into a runtime JSON timeline.

Reads:
    frontend/public/Zumba_Mappings_AllSongs_Adaptive_5Modes_18Steps.xlsx
Writes:
    frontend/public/zumba-mappings-18steps.json

The browser must never parse the Excel file. This script is the single source of
truth that turns the workbook into the lightweight timeline JSON consumed by the
Zumba page. It fails loudly when the workbook and the GLB assets disagree.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
PUBLIC_DIR = REPO_ROOT / "frontend" / "public"
WORKBOOK_PATH = PUBLIC_DIR / "Zumba_Mappings_AllSongs_Adaptive_5Modes_18Steps.xlsx"
OUTPUT_PATH = PUBLIC_DIR / "zumba-mappings-18steps.json"
ASSET_ROOT = PUBLIC_DIR / "zumba pose"

VERSION = "1.1.0"
MODES = ["Beginner", "Easy", "Moderate", "High", "HIIT"]

# Master move table. Keyed by the exact "Move Name" used in the workbook.
#   key        -> stable, URL-safe key used by the frontend manifest + GLB folders
#   folder     -> asset folder under /zumba pose/
#   in/main/out-> exact GLB filenames inside that folder
#   backendKey -> key the backend's reference_angles expects (target_move on the wire)
MOVE_TABLE = {
    "Body Rolls": {
        "key": "body_rolls", "backendKey": "body_rolls", "folder": "Body Rolls",
        "in": "Idle in Body Rolls_compressed.glb",
        "main": "Main Body Rolls_compressed.glb",
        "out": "Idle out Body Rolls_compressed.glb",
    },
    "Cumbia Step": {
        "key": "cumbia_step", "backendKey": "cumbia_step", "folder": "Cumbia Step",
        "in": "Idle In Cumbia Step_compressed.glb",
        "main": "Main Cumbia Step_compressed.glb",
        "out": "Idle Out Cumbia Step_compressed.glb",
    },
    "Grape Vine": {
        "key": "grape_vine", "backendKey": "grape_vine", "folder": "Grape Vine",
        "in": "Idle In Grape Vine_compressed.glb",
        "main": "Main Grape Vine_compressed.glb",
        "out": "Idle Out Grape Vine_compressed.glb",
    },
    "Heel Taps": {
        "key": "heel_taps", "backendKey": "heel_taps", "folder": "Heel Taps",
        "in": "Idle in Heel Taps_compressed.glb",
        "main": "Main Heel Taps_compressed.glb",
        "out": "Idle out Heel Taps_compressed.glb",
    },
    "Jumping Jacks": {
        "key": "jumping_jacks", "backendKey": "jumping_jacks", "folder": "Jumping Jacks",
        "in": "Idle In Jumping Jacks_compressed.glb",
        "main": "Main Jumping Jacks_compressed.glb",
        "out": "Idle Out Jumping Jacks_compressed.glb",
    },
    "Knee Lifts": {
        "key": "knee_lifts", "backendKey": "knee_lifts", "folder": "Knee Lifts",
        "in": "Idle in Knee Lift_compressed.glb",
        "main": "Main Knee Lift_compressed.glb",
        "out": "Idle out Knee Lift_compressed.glb",
    },
    "Lunge with Punches": {
        "key": "lunge_with_punches", "backendKey": "lunge_with_punches", "folder": "Lunge with Punches",
        "in": "Idle in Lunge with Punches_compressed.glb",
        "main": "Main Lunge with Punches_compressed.glb",
        "out": "Idle out Lunge with Punches_compressed.glb",
    },
    "Mambo": {
        "key": "mambo", "backendKey": "mambo_step", "folder": "Mambo",
        "in": "Idle in Mambo_compressed.glb",
        "main": "Main Mambo_compressed.glb",
        "out": "Idle out Mambo_compressed.glb",
    },
    "March in Place": {
        "key": "march_in_place", "backendKey": "march_in_place", "folder": "March in Place",
        "in": "Idle in March in place_compressed.glb",
        "main": "Main March in place_compressed.glb",
        "out": "Idle out March in place_compressed.glb",
    },
    "Merengue March": {
        "key": "merengue_march", "backendKey": "merengue_march", "folder": "Merengue March",
        "in": "Idle In Merengue March_compressed.glb",
        "main": "Main Merengue March_compressed.glb",
        "out": "Idle Out Merengue March_compressed.glb",
    },
    "Reggaeton Stomp": {
        "key": "reggaeton_stomp", "backendKey": "reggaeton_stomp", "folder": "Reggaeton Stomp",
        "in": "Idle in Reggaeton stomp_compressed.glb",
        "main": "Main Reggaeton stomp_compressed.glb",
        "out": "Idle out Reggaeton stomp_compressed.glb",
    },
    "Shimmy": {
        "key": "shimmy", "backendKey": "shimmy", "folder": "Shimmy",
        "in": "Idle in Shimmy_compressed.glb",
        "main": "Main Shimmy_compressed.glb",
        "out": "Idle out Shimmy_compressed.glb",
    },
    "Side Punches": {
        "key": "side_punches", "backendKey": "side_punches", "folder": "Side Punches",
        "in": "Idle in Side punches_compressed.glb",
        "main": "Main Side punches_compressed.glb",
        "out": "Idle out Side punches_compressed.glb",
    },
    "Squat With Clap": {
        "key": "squat_with_clap", "backendKey": "squat_with_clap", "folder": "Squat With Clap",
        "in": "Idle In Squat With Clap_compressed.glb",
        "main": "Main Squat With Clap_compressed.glb",
        "out": "Idle Out Squat With Clap_compressed.glb",
    },
    "Step Clap": {
        "key": "step_clap", "backendKey": "step_clap", "folder": "Step Clap",
        "in": "Idle in Step clap_compressed.glb",
        "main": "Main Step clap_compressed.glb",
        "out": "Idle out Step clap_compressed.glb",
    },
    "Step Touch": {
        "key": "step_touch", "backendKey": "step_touch", "folder": "Step Touch",
        "in": "Idle in Step touch_compressed.glb",
        "main": "Main Step touch_compressed.glb",
        "out": "Idle out Step touch_compressed.glb",
    },
    "Twist Step": {
        "key": "twist_step", "backendKey": "twist_step", "folder": "Twist Step",
        "in": "Idle in Twist Step_compressed.glb",
        "main": "Main Twist Step_compressed.glb",
        "out": "Idle out Twist Step_compressed.glb",
    },
    "Zumba Turn(Pivot)": {
        "key": "zumba_turn_pivot", "backendKey": "zumba_turn_(pivot)", "folder": "Zumba Turn(Pivot)",
        "in": "Idle in Zumba turn (pivot)_compressed.glb",
        "main": "Main Zumba turn (pivot)_compressed.glb",
        "out": "Idle out Zumba turn (pivot)_compressed.glb",
    },
}

# Client-corrected choreography sheets (beat-aligned pilot format).
# Key: (song, mode) -> corrected workbook + sheet. As the client delivers more
# corrected sheets, add entries here — no other converter change needed.
CORRECTED_SHEETS = {
    ("All Night", "High"): {
        "file": REPO_ROOT / "docs" / "Zumba_AllNight_High_Pilot_Corrected.xlsx",
        "sheet": "Pilot Corrected",
    },
}

# Extra columns present only in the corrected-sheet format.
CORRECTED_EXTRA_COLUMNS = [
    "Beat Phase Offset (s)",
    "GLB Main Duration (s)",
    "Planned Loops in Group",
    "Animation Speed Scale",
]

# Safety band for playback speed adjustment; outside this the animation looks
# visibly wrong, so the converter refuses the sheet.
SPEED_SCALE_HARD_MIN = 0.45
SPEED_SCALE_HARD_MAX = 1.15
# Beyond this we only warn (sheet value kept, but flagged for the client).
SPEED_SCALE_WARN_DELTA = 0.02
GLB_DURATION_WARN_DELTA_S = 0.05

# Song audio (the Zumba session clock + the music the user dances to).
SONG_AUDIO = {
    "All Night": "Taketones_AllNight.wav",
    "Beach Dance": "Taketones_BeachDance.wav",
    "Blue Print": "Taketones_BluePrint.wav",
    "Bollywood": "Taketones_Bollywood.wav",
    "Indian Energy": "Taketones_IndianEnergy.wav",
    "Like a Dynamite": "Taketones_LikeA Dynamite.wav",
}
AUDIO_ROOT = PUBLIC_DIR / "zumba-tones"

EXPECTED_COLUMNS = [
    "Song Title", "Artist", "Mode", "BPM (est)", "Block Size (counts)",
    "Start Time (mm:ss.ms)", "End Time (mm:ss.ms)", "Duration (s)",
    "Move Group Index", "Move Group Length (blocks)", "Move ID", "Move Name",
    "Side (L/R)", "Coach Cue",
]


class BuildError(Exception):
    pass


def fail(msg: str) -> "BuildError":
    return BuildError(msg)


def asset_path(folder: str, filename: str) -> str:
    return f"/zumba pose/{folder}/{filename}"


def parse_time_ms(raw, sheet: str, row_idx: int, label: str) -> int:
    """Parse 'mm:ss.ms' (or seconds) into integer milliseconds."""
    if raw is None:
        raise fail(f"{sheet} row {row_idx}: missing {label}")
    if isinstance(raw, (int, float)):
        # Stored as a number of seconds.
        return int(round(float(raw) * 1000))
    text = str(raw).strip()
    m = re.match(r"^(\d+):(\d{1,2})(?:\.(\d{1,3}))?$", text)
    if not m:
        m2 = re.match(r"^(\d+(?:\.\d+)?)$", text)
        if m2:
            return int(round(float(m2.group(1)) * 1000))
        raise fail(f"{sheet} row {row_idx}: invalid {label} '{text}'")
    minutes = int(m.group(1))
    seconds = int(m.group(2))
    millis = int((m.group(3) or "0").ljust(3, "0"))
    return (minutes * 60 + seconds) * 1000 + millis


def build_moves() -> dict:
    """Validate every move folder + Main GLB exists and emit the moves map."""
    moves: dict = {}
    for name, info in MOVE_TABLE.items():
        folder_dir = ASSET_ROOT / info["folder"]
        if not folder_dir.is_dir():
            raise fail(f"Move '{name}' has no asset folder: {folder_dir}")
        for phase in ("in", "main", "out"):
            file_path = folder_dir / info[phase]
            if not file_path.is_file():
                if phase == "main":
                    raise fail(f"Move '{name}' is missing its Main GLB: {file_path}")
                raise fail(f"Move '{name}' is missing its {phase} GLB: {file_path}")
        moves[info["key"]] = {
            "id": 0,  # filled in from the Moves sheet
            "key": info["key"],
            "backendKey": info["backendKey"],
            "name": name,
            "folder": info["folder"],
            "in": asset_path(info["folder"], info["in"]),
            "main": asset_path(info["folder"], info["main"]),
            "out": asset_path(info["folder"], info["out"]),
        }
    return moves


def read_moves_sheet(wb) -> dict:
    """Return {move_id: move_name} and confirm every move maps to MOVE_TABLE."""
    ws = wb["Moves"]
    id_to_name: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        move_id, move_name = row[0], row[1]
        if move_name is None:
            continue
        name = str(move_name).strip()
        if name not in MOVE_TABLE:
            raise fail(f"Moves sheet move '{name}' has no asset mapping in MOVE_TABLE")
        id_to_name[int(move_id)] = name
    if len(id_to_name) != 18:
        raise fail(f"Expected 18 moves in Moves sheet, found {len(id_to_name)}")
    return id_to_name


def parse_mode(raw: str, sheet: str) -> str:
    mode = str(raw).strip()
    if mode not in MODES:
        raise fail(f"{sheet}: unexpected mode '{mode}'")
    return mode


def clean_cue(raw) -> str:
    if raw is None:
        return ""
    text = str(raw)
    # The workbook stores an en-dash that round-trips as U+FFFD in some exports.
    text = text.replace("�", "-")
    return text.strip()


def build_timeline_sheet(wb, sheet: str, id_to_name: dict) -> list:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise fail(f"{sheet}: empty sheet")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    for col in EXPECTED_COLUMNS:
        if col not in header:
            raise fail(f"{sheet}: missing required column '{col}'")
    col = {name: header.index(name) for name in EXPECTED_COLUMNS}

    blocks = []
    prev_end = None
    for idx, raw in enumerate(rows[1:], start=2):
        if not raw or raw[col["Song Title"]] is None:
            continue

        move_id = raw[col["Move ID"]]
        if move_id is None or int(move_id) not in id_to_name:
            raise fail(f"{sheet} row {idx}: Move ID '{move_id}' not in Moves sheet")
        move_name = id_to_name[int(move_id)]
        sheet_move_name = str(raw[col["Move Name"]]).strip() if raw[col["Move Name"]] else move_name
        if sheet_move_name != move_name:
            raise fail(
                f"{sheet} row {idx}: Move ID {move_id} maps to '{move_name}' "
                f"but row says '{sheet_move_name}'"
            )
        info = MOVE_TABLE[move_name]

        start_ms = parse_time_ms(raw[col["Start Time (mm:ss.ms)"]], sheet, idx, "Start Time")
        end_ms = parse_time_ms(raw[col["End Time (mm:ss.ms)"]], sheet, idx, "End Time")
        if end_ms <= start_ms:
            raise fail(f"{sheet} row {idx}: End ({end_ms}) <= Start ({start_ms})")
        if prev_end is not None and start_ms < prev_end:
            raise fail(
                f"{sheet} row {idx}: time overlap start={start_ms} < prev_end={prev_end}"
            )
        prev_end = end_ms

        side_raw = raw[col["Side (L/R)"]]
        side = None
        if side_raw is not None:
            s = str(side_raw).strip().upper()
            if s in ("L", "R"):
                side = s
            elif s:
                raise fail(f"{sheet} row {idx}: invalid Side '{side_raw}'")

        blocks.append({
            "songTitle": str(raw[col["Song Title"]]).strip(),
            "artist": str(raw[col["Artist"]]).strip() if raw[col["Artist"]] else "",
            "mode": parse_mode(raw[col["Mode"]], sheet),
            "bpm": float(raw[col["BPM (est)"]]) if raw[col["BPM (est)"]] is not None else 0.0,
            "blockSizeCounts": int(raw[col["Block Size (counts)"]] or 8),
            "startMs": start_ms,
            "endMs": end_ms,
            "durationMs": end_ms - start_ms,
            "moveGroupIndex": int(raw[col["Move Group Index"]] or 0),
            "moveGroupLengthBlocks": int(raw[col["Move Group Length (blocks)"]] or 0),
            "moveId": int(move_id),
            "moveKey": info["key"],
            "moveName": move_name,
            "side": side,
            "coachCue": clean_cue(raw[col["Coach Cue"]]),
        })

    if not blocks:
        raise fail(f"{sheet}: no timeline rows")
    return blocks


def read_glb_animation_duration(path: Path) -> float:
    """Read the longest animation duration (seconds) straight from a GLB file.

    GLB layout: 12-byte header, then chunks. Chunk 0 is the JSON scene
    description; each animation sampler's *input* accessor holds keyframe
    times, and its `max[0]` is the clip end time. This lets the converter
    validate the client's hand-filled 'GLB Main Duration (s)' column against
    the real asset without any GLTF library.
    """
    import struct

    with open(path, "rb") as fh:
        header = fh.read(12)
        if len(header) < 12 or header[:4] != b"glTF":
            raise fail(f"Not a GLB file: {path}")
        chunk_len, chunk_type = struct.unpack("<II", fh.read(8))
        if chunk_type != 0x4E4F534A:  # 'JSON'
            raise fail(f"GLB missing JSON chunk: {path}")
        doc = json.loads(fh.read(chunk_len).decode("utf-8"))

    accessors = doc.get("accessors", [])
    duration = 0.0
    for anim in doc.get("animations", []):
        for sampler in anim.get("samplers", []):
            idx = sampler.get("input")
            if idx is None or idx >= len(accessors):
                continue
            acc_max = accessors[idx].get("max")
            if acc_max:
                duration = max(duration, float(acc_max[0]))
    return duration


def build_corrected_timeline(song: str, mode: str, spec: dict, id_to_name: dict):
    """Parse a client-corrected sheet into (blocks, meta, warnings).

    Corrected sheets add: per-row beat phase offset, GLB main duration,
    planned loops, animation speed scale, plus two special non-move rows —
    a lead-in hold at the start and an end hold for the audio tail.
    """
    path = spec["file"]
    if not path.is_file():
        raise fail(f"Corrected sheet for {song} - {mode} not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = spec["sheet"]
    if sheet not in wb.sheetnames:
        raise fail(f"Corrected workbook {path.name} has no sheet '{sheet}'")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise fail(f"{path.name}/{sheet}: empty sheet")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    required = [c for c in EXPECTED_COLUMNS if c != "Block Size (counts)"] + CORRECTED_EXTRA_COLUMNS
    for col_name in required + ["Block Size (counts)"]:
        if col_name not in header:
            raise fail(f"{path.name}/{sheet}: missing required column '{col_name}'")
    col = {name: header.index(name) for name in header if name}

    blocks = []
    warnings: list = []
    meta = {
        "corrected": True,
        "beatPhaseOffsetMs": 0,
        "leadInEndMs": None,
        "outroStartMs": None,
        "outroEndMs": None,
    }
    prev_end = None

    for idx, raw in enumerate(rows[1:], start=2):
        if not raw or raw[col["Song Title"]] is None:
            continue

        start_ms = parse_time_ms(raw[col["Start Time (mm:ss.ms)"]], sheet, idx, "Start Time")
        end_ms = parse_time_ms(raw[col["End Time (mm:ss.ms)"]], sheet, idx, "End Time")
        if end_ms <= start_ms:
            raise fail(f"{sheet} row {idx}: End ({end_ms}) <= Start ({start_ms})")
        if prev_end is not None and start_ms < prev_end:
            raise fail(f"{sheet} row {idx}: time overlap start={start_ms} < prev_end={prev_end}")
        prev_end = end_ms

        offset_raw = raw[col["Beat Phase Offset (s)"]]
        if offset_raw is not None:
            meta["beatPhaseOffsetMs"] = int(round(float(offset_raw) * 1000))

        move_id = int(raw[col["Move ID"]] or 0)
        group_index = int(raw[col["Move Group Index"]] if raw[col["Move Group Index"]] is not None else 0)
        row_name = str(raw[col["Move Name"]] or "").strip()

        # Special hold rows: lead-in before the first beat, end hold on the tail.
        if move_id == 0 or move_id not in id_to_name:
            lowered = row_name.lower()
            if group_index == -1 or "lead-in" in lowered or "lead in" in lowered:
                meta["leadInEndMs"] = end_ms
                continue
            if "idle out" in lowered or "end hold" in lowered or "outro" in lowered:
                meta["outroStartMs"] = start_ms
                meta["outroEndMs"] = end_ms
                continue
            raise fail(f"{sheet} row {idx}: unknown special row '{row_name}' (Move ID {move_id})")

        move_name = id_to_name[move_id]
        if row_name and row_name != move_name:
            raise fail(f"{sheet} row {idx}: Move ID {move_id} maps to '{move_name}' but row says '{row_name}'")
        info = MOVE_TABLE[move_name]

        side_raw = raw[col["Side (L/R)"]]
        side = None
        if side_raw is not None:
            s = str(side_raw).strip().upper()
            if s in ("L", "R"):
                side = s
            elif s:
                raise fail(f"{sheet} row {idx}: invalid Side '{side_raw}'")

        glb_duration = float(raw[col["GLB Main Duration (s)"]] or 0)
        planned_loops = int(raw[col["Planned Loops in Group"]] or 0)
        speed_scale = float(raw[col["Animation Speed Scale"]] or 1.0)
        if not (SPEED_SCALE_HARD_MIN <= speed_scale <= SPEED_SCALE_HARD_MAX):
            raise fail(
                f"{sheet} row {idx}: Animation Speed Scale {speed_scale} outside safe band "
                f"[{SPEED_SCALE_HARD_MIN}, {SPEED_SCALE_HARD_MAX}]"
            )

        blocks.append({
            "songTitle": str(raw[col["Song Title"]]).strip(),
            "artist": str(raw[col["Artist"]]).strip() if raw[col["Artist"]] else "",
            "mode": parse_mode(raw[col["Mode"]], sheet),
            "bpm": float(raw[col["BPM (est)"]]) if raw[col["BPM (est)"]] is not None else 0.0,
            "blockSizeCounts": int(raw[col["Block Size (counts)"]] or 8),
            "startMs": start_ms,
            "endMs": end_ms,
            "durationMs": end_ms - start_ms,
            "moveGroupIndex": group_index,
            "moveGroupLengthBlocks": int(raw[col["Move Group Length (blocks)"]] or 0),
            "moveId": move_id,
            "moveKey": info["key"],
            "moveName": move_name,
            "side": side,
            "coachCue": clean_cue(raw[col["Coach Cue"]]),
            "glbMainDurationSec": glb_duration,
            "plannedLoops": planned_loops,
            "animationSpeedScale": speed_scale,
        })

    if not blocks:
        raise fail(f"{path.name}/{sheet}: no move rows")

    # --- Cross-check the client's numbers against the math and the real GLBs ---
    groups: dict = {}
    for b in blocks:
        groups.setdefault(b["moveGroupIndex"], []).append(b)
    for gid, gblocks in sorted(groups.items()):
        first = gblocks[0]
        group_sec = (gblocks[-1]["endMs"] - gblocks[0]["startMs"]) / 1000.0
        sheet_glb = first["glbMainDurationSec"]
        sheet_loops = first["plannedLoops"]
        sheet_scale = first["animationSpeedScale"]

        # 1) Sheet's GLB duration vs the actual asset file.
        glb_path = ASSET_ROOT / MOVE_TABLE[first["moveName"]]["folder"] / MOVE_TABLE[first["moveName"]]["main"]
        measured = read_glb_animation_duration(glb_path)
        if measured > 0 and abs(measured - sheet_glb) > GLB_DURATION_WARN_DELTA_S:
            warnings.append(
                f"group {gid} ({first['moveName']}): sheet GLB duration {sheet_glb:.3f}s "
                f"!= measured {measured:.3f}s ({glb_path.name})"
            )

        # 2) Loop math: loops*glb/scale should fill the group window.
        if sheet_glb > 0 and group_sec > 0:
            expected_loops = max(1, round(group_sec / sheet_glb))
            expected_scale = (expected_loops * sheet_glb) / group_sec
            if sheet_loops and sheet_loops != expected_loops:
                warnings.append(
                    f"group {gid} ({first['moveName']}): sheet loops {sheet_loops} "
                    f"!= recomputed {expected_loops} (group {group_sec:.3f}s / glb {sheet_glb:.3f}s)"
                )
            if abs(sheet_scale - expected_scale) > SPEED_SCALE_WARN_DELTA:
                warnings.append(
                    f"group {gid} ({first['moveName']}): sheet speed scale {sheet_scale:.3f} "
                    f"!= recomputed {expected_scale:.3f}"
                )

    if meta["leadInEndMs"] is None and meta["beatPhaseOffsetMs"]:
        meta["leadInEndMs"] = meta["beatPhaseOffsetMs"]
    return blocks, meta, warnings


def main() -> int:
    if not WORKBOOK_PATH.is_file():
        print(f"ERROR: workbook not found at {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        moves = build_moves()
        id_to_name = read_moves_sheet(wb)
        for move_id, name in id_to_name.items():
            moves[MOVE_TABLE[name]["key"]]["id"] = move_id

        songs: dict = {}
        timeline_sheets = 0
        total_rows = 0
        for sheet in wb.sheetnames:
            if sheet in ("Moves", "18 Step Mapping Rules"):
                continue
            if " - " not in sheet:
                raise fail(f"Unexpected sheet name '{sheet}'")
            song_title, mode = sheet.rsplit(" - ", 1)
            song_title = song_title.strip()
            mode = mode.strip()
            if mode not in MODES:
                raise fail(f"Sheet '{sheet}' has unknown mode '{mode}'")
            blocks = build_timeline_sheet(wb, sheet, id_to_name)
            songs.setdefault(song_title, {})[mode] = blocks
            timeline_sheets += 1
            total_rows += len(blocks)

        # Apply client-corrected sheets on top of the base workbook. Corrected
        # timelines replace the base blocks and gain beat-alignment metadata.
        timeline_meta: dict = {}
        corrected_count = 0
        all_warnings: list = []
        for (c_song, c_mode), spec in CORRECTED_SHEETS.items():
            if c_song not in songs or c_mode not in songs[c_song]:
                raise fail(f"Corrected sheet targets unknown timeline: {c_song} - {c_mode}")
            c_blocks, c_meta, c_warnings = build_corrected_timeline(c_song, c_mode, spec, id_to_name)
            songs[c_song][c_mode] = c_blocks
            timeline_meta.setdefault(c_song, {})[c_mode] = c_meta
            corrected_count += 1
            all_warnings.extend(f"[{c_song} - {c_mode}] {w}" for w in c_warnings)
        # Non-corrected timelines get explicit default meta so the frontend
        # can branch without guessing.
        for song, modes_map in songs.items():
            for mode in modes_map:
                if mode not in timeline_meta.get(song, {}):
                    timeline_meta.setdefault(song, {})[mode] = {
                        "corrected": False,
                        "beatPhaseOffsetMs": 0,
                        "leadInEndMs": None,
                        "outroStartMs": None,
                        "outroEndMs": None,
                    }

        # Structural validation.
        if len(songs) != 6:
            raise fail(f"Expected 6 songs, found {len(songs)}: {list(songs)}")
        for song, modes in songs.items():
            missing = [m for m in MODES if m not in modes]
            if missing:
                raise fail(f"Song '{song}' missing modes: {missing}")
        used_keys = {b["moveKey"] for modes in songs.values() for blocks in modes.values() for b in blocks}
        missing_assets = [k for k in used_keys if k not in moves]
        if missing_assets:
            raise fail(f"Timeline references moves with no asset: {missing_assets}")

        # Resolve + validate one audio track per song.
        audio = {}
        for song in songs:
            filename = SONG_AUDIO.get(song)
            if not filename:
                raise fail(f"Song '{song}' has no audio mapping in SONG_AUDIO")
            if not (AUDIO_ROOT / filename).is_file():
                raise fail(f"Song '{song}' audio file missing: {AUDIO_ROOT / filename}")
            audio[song] = f"/zumba-tones/{filename}"

    except BuildError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print("Asset validation: FAILED", file=sys.stderr)
        return 1

    output = {
        "version": VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "modes": MODES,
        "moves": moves,
        "audio": audio,
        "timelineMeta": timeline_meta,
        "songs": songs,
    }
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Songs: {len(songs)}")
    print(f"Modes: {len(MODES)}")
    print(f"Timeline sheets: {timeline_sheets}")
    print(f"Timeline rows: {total_rows}")
    print(f"Corrected timelines: {corrected_count}")
    if all_warnings:
        print(f"Sheet cross-check warnings: {len(all_warnings)}")
        for w in all_warnings:
            print(f"  WARN: {w}")
    else:
        print("Sheet cross-check: clean")
    print(f"Moves: {len(moves)}")
    print("Asset validation: passed")
    print(f"Wrote: {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
